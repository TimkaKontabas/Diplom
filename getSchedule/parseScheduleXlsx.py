import openpyxl
from pydantic import BaseModel
from typing import List


class Lesson(BaseModel):
    day: int
    name: str
    location: str

class GroupData(BaseModel):
    lessons: List[Lesson]

class Schedule(BaseModel):
    groups: List[int]
    dates: List[str]
    data: List[GroupData]


def getSheduleSheet():
    wb = openpyxl.load_workbook('schedule.xlsx')
    return wb['Лист2']
    

def parseScheduleXlsx() -> Schedule:
    sheet = getSheduleSheet()
    shedule = Schedule(groups=[], dates=[], data=[])
    for row in sheet.iter_rows(min_row=2):
        findGroupInRow(row, shedule)

    return shedule

def isCellGroup(cell):
    if type(cell.value) is str and 'группа' in cell.value:
        return True

def isRowGroup(row):
    for cell in row:
        if isCellGroup(cell):
            return True
    return False

def findGroupInRow(row, shedule):
    if isRowGroup(row):
        addGroup(row, shedule)



def addGroup(row, shedule):
    for cell in row:
        if isCellGroup(cell):
            shedule.groups.append(int(cell.value.split()[0]))

def printSheet(start:int, rows:int):
    sheet = getSheduleSheet()
    index_row = 0
    with open("output.txt", "w") as f:
        for row in sheet.iter_rows(min_row=start, max_row=start+rows):

            print([cell.value for cell in row[:-4]], file=f)
            index_row += 1
    
printSheet(2, 50)
# shedule = parseScheduleXlsx()
# print(shedule.model_dump_json())